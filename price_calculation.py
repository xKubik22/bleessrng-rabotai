import openpyxl
from works_indexes import indexes 

def max_rows(name_sheet):
	wb_r = openpyxl.load_workbook('ZP.xlsx')
	ws_r = wb_r[name_sheet]
	max_rows = ws_r.max_row
	print(max_rows)
	return max_rows

def read_price_list():
	wb_r = openpyxl.load_workbook('ZP.xlsx')
	ws_r = wb_r.active
	read_price_list = []
	price_list_done = {}
	if 'Цены' in wb_r.sheetnames:
		wb_r.active = wb_r['Цены']
		for row in ws_r.iter_rows(min_row= 0, max_row=ws_r.max_row-1, max_col=ws_r.max_column, min_col=1):
			for cell in row:
				read_price_list.append(cell.value)	
	else:
		wb_r.create_sheet("Цены")
	price_list_done = {read_price_list[i]: read_price_list[i+1] for i in range(0, len(read_price_list) - 1,2)}
	price_list_done.pop(None)
	wb_r.close()
	return price_list_done
	
def calculate_price(object_work):
	price_list = read_price_list()
	# print(price_list)
	total_price = 0
	for key in price_list.keys():
		total_price += int(price_list[key]) * object_work[indexes[key]]
	return total_price

if __name__ == "__main__":
	wb_r = openpyxl.load_workbook('ZP.xlsx')
	ws_r = wb_r.active
	max_rows('Расчет')
	read_price_list()
	calculate_price([0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])
	wb_r.save('ZP.xlsx')