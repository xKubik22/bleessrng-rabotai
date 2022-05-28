import openpyxl
from config import Config
from works_indexes import indexes


class WorksAdder:
	PATH = None

	def __init__(self, obj):
		self.init_path()
		self.price_list = self.get_price_list()
		self.works = obj.get_works()
		price = self.calculate_price()
		works_info = [obj.get_name()]
		works_info.extend(self.works)
		works_info.append(price)
		self.write_object_works(works_info)
		self.__del__()

	@classmethod
	def init_path(cls):
		cls.PATH = Config.get_calc_path()

	def __del__(self):
		pass

	@classmethod
	def check_list(cls, workbook, list_name: str) -> bool:
		if list_name not in workbook.sheetnames:
			workbook.create_sheet(list_name)
			workbook.save(cls.PATH)
			print(f'В файле {cls.PATH} отсутсвовал лист {list_name}, который нужно заполнить')
			return False

		return True

	@classmethod
	def get_price_list(cls) -> dict:
		workbook = cls.get_workbook()
		price_list = []
		if not cls.check_list(workbook, 'Цены'):
			return {}

		worksheet = workbook['Цены']
		for row in worksheet.iter_rows(min_row=0, max_row=worksheet.max_row - 1, max_col=worksheet.max_column, min_col=1):
			for cell in row:
				price_list.append(cell.value)

		price_list_done = {price_list[i]: price_list[i + 1] for i in range(0, len(price_list) - 1, 2)}

		if len(price_list_done) == 0:
			return price_list_done
		try:
			price_list_done.pop(None)
			workbook.close()
		finally:
			return price_list_done

	def calculate_price(self) -> int:
		total_price = 0
		if len(self.price_list) == 0:
			print(f'Необходимо заполнить лист "Цены" в файле {WorksAdder.PATH}')
			self.__del__()

		for key in self.price_list.keys():
			total_price += int(self.price_list[key]) * self.works[indexes[key]]

		return total_price

	@classmethod
	def get_workbook(cls):
		try:
			workbook = openpyxl.load_workbook(cls.PATH)
		except FileNotFoundError:
			workbook = openpyxl.Workbook()
			workbook.save(cls.PATH)

		return workbook

	@classmethod
	def write_object_works(cls, object_info):
		workbook = cls.get_workbook()
		if cls.check_list(workbook, 'Расчет'):
			worksheet = workbook['Расчет']
			worksheet.cell(row=worksheet.max_row, column=1)
			worksheet.append(object_info)
			workbook.save(cls.PATH)
			workbook.close()



if __name__ == "__main__":
	object_info = ['какая-то херь', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 100]
	write_object_works(object_info)

